# ruff: noqa: E501

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from uuid import UUID

import openpyxl

from app.modules.imports.repository import ImportRepository


class StudentImportValidator:
    def __init__(self, repository: ImportRepository, tenant_id: UUID, context_branch_id: UUID | None = None) -> None:
        self.repository = repository
        self.tenant_id = tenant_id
        self.context_branch_id = context_branch_id
        self.required_columns = {
            "admission_number": ["Admission No", "Admission Number"],
            "student_name": ["Student Name", "Student Full Name"],
            "gender": ["Gender"],
            "date_of_birth": ["Date Of Birth", "Date of Birth"],
            "academic_year": ["Academic Year"],
            "programme": ["Programme / Stream"],
            "year_level": ["Year Level"],
            "batch": ["Batch"],
            "section": ["Section"],
            "joining_date": ["Joining Date"],
            "guardian_name": ["Guardian Name"],
            "guardian_relationship": ["Relationship", "Guardian Relationship"],
            "guardian_mobile": ["Guardian Phone", "Guardian Mobile"],
        }
        self.optional_columns = {
            "branch_code": ["Branch Code"],
            "student_mobile": ["Student Mobile"],
            "student_email": ["Student Email"],
            "roll_number": ["Roll No", "Roll Number"],
            "ending_date": ["Ending Date"],
            "guardian_email": ["Guardian Email"],
            "student_created": ["Student Created"],
        }

    def _header_map(self, headers: tuple[Any, ...]) -> dict[str, int]:
        return {str(header).strip(): index for index, header in enumerate(headers) if header is not None and str(header).strip()}

    def _find_column(self, header_map: dict[str, int], aliases: list[str]) -> int | None:
        normalized = {header.lower(): index for header, index in header_map.items()}
        for alias in aliases:
            index = normalized.get(alias.lower())
            if index is not None:
                return index
        return None

    def _value(self, row_values: tuple[Any, ...], columns: dict[str, int | None], key: str) -> Any:
        index = columns.get(key)
        if index is None or index >= len(row_values):
            return None
        return row_values[index]

    def _clean_string(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, Decimal):
            return str(value.normalize())
        return str(value).strip()

    def _parse_date(self, value: Any) -> str | None:
        """Accept Excel date cells or day-first text, then normalize for database storage."""
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = self._clean_string(value)
        for date_format in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
            try:
                return datetime.strptime(text, date_format).date().isoformat()
            except ValueError:
                continue
        return None

    def _normalize_year_level(self, value: str) -> str:
        normalized = value.strip().lower().replace("_", " ").replace("-", " ")
        if normalized in {"1", "first", "first year", "year 1", "fy"}:
            return "1"
        if normalized in {"2", "second", "second year", "year 2", "sy"}:
            return "2"
        return ""

    def parse_and_validate(self, file_content: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active
        if not sheet:
            raise ValueError("No active sheet found in Excel file.")

        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            raise ValueError("Excel file is empty.")

        header_map = self._header_map(headers)
        columns = {
            key: self._find_column(header_map, aliases)
            for key, aliases in {**self.required_columns, **self.optional_columns}.items()
        }
        missing_headers = [
            aliases[0]
            for key, aliases in self.required_columns.items()
            if columns[key] is None
        ]
        if missing_headers:
            raise ValueError(f"Missing required columns: {', '.join(missing_headers)}")

        results = []
        summary = {"total_rows": 0, "valid_rows": 0, "warning_rows": 0, "rejected_rows": 0}

        seen_admission_numbers: set[tuple[str, str]] = set()
        seen_roll_numbers: set[tuple[str, str, str]] = set()

        for row_idx, row_values in enumerate(rows_iter, start=2):
            # Skip completely empty rows
            if not any(row_values):
                continue
            if row_idx == 2:
                first_value = self._clean_string(row_values[0] if row_values else "")
                if first_value.lower() in {"required", "optional"}:
                    continue
            
            summary["total_rows"] += 1

            raw_data = {
                header: row_values[idx] for header, idx in header_map.items()
            }
            normalized_data = {}
            errors = []
            status = "VALID"

            branch_code = self._clean_string(self._value(row_values, columns, "branch_code"))
            admission_number = self._clean_string(self._value(row_values, columns, "admission_number"))
            student_name = self._clean_string(self._value(row_values, columns, "student_name"))
            dob = self._value(row_values, columns, "date_of_birth")
            gender = self._clean_string(self._value(row_values, columns, "gender")).upper()
            student_mobile = self._clean_string(self._value(row_values, columns, "student_mobile"))
            student_email = self._clean_string(self._value(row_values, columns, "student_email"))
            academic_year_str = self._clean_string(self._value(row_values, columns, "academic_year"))
            programme_str = self._clean_string(self._value(row_values, columns, "programme"))
            year_level_label = self._clean_string(self._value(row_values, columns, "year_level"))
            batch_str = self._clean_string(self._value(row_values, columns, "batch"))
            section_str = self._clean_string(self._value(row_values, columns, "section"))
            roll_number = self._clean_string(self._value(row_values, columns, "roll_number"))
            joining_date = self._value(row_values, columns, "joining_date")
            ending_date = self._value(row_values, columns, "ending_date")
            guardian_name = self._clean_string(self._value(row_values, columns, "guardian_name"))
            guardian_rel = self._clean_string(self._value(row_values, columns, "guardian_relationship")).upper()
            guardian_mobile = self._clean_string(self._value(row_values, columns, "guardian_mobile"))
            guardian_email = self._clean_string(self._value(row_values, columns, "guardian_email"))

            # Date parsing
            dob_val = self._parse_date(dob)
            if dob_val is None:
                errors.append({"field": "Date of Birth", "message": "Invalid date format. Use DD/MM/YYYY or DD-MM-YYYY."})
                status = "REJECTED"

            joining_val = self._parse_date(joining_date)
            if joining_val is None:
                errors.append({"field": "Joining Date", "message": "Invalid date format. Use DD/MM/YYYY or DD-MM-YYYY."})
                status = "REJECTED"

            ending_val = self._parse_date(ending_date)
            if ending_date not in (None, "") and ending_val is None:
                errors.append({"field": "Ending Date", "message": "Invalid date format. Use DD/MM/YYYY or DD-MM-YYYY."})
                status = "REJECTED"

            # Required fields check
            if not admission_number:
                errors.append({"field": "Admission Number", "message": "Required field missing."})
                status = "REJECTED"
            if not student_name:
                errors.append({"field": "Student Full Name", "message": "Required field missing."})
                status = "REJECTED"
            if not guardian_name:
                errors.append({"field": "Guardian Name", "message": "Required field missing."})
                status = "REJECTED"
            if not guardian_mobile:
                errors.append({"field": "Guardian Mobile", "message": "Required field missing."})
                status = "REJECTED"
            if not year_level_label:
                errors.append({"field": "Year Level", "message": "Required field missing."})
                status = "REJECTED"
            if not batch_str:
                errors.append({"field": "Batch", "message": "Required field missing."})
                status = "REJECTED"
            if gender not in ["MALE", "FEMALE", "OTHER"]:
                errors.append({"field": "Gender", "message": "Must be one of MALE, FEMALE, OTHER."})
                status = "REJECTED"
            
            # Guardian Relationship validation
            valid_relations = ["FATHER", "MOTHER", "LEGAL_GUARDIAN", "RELATIVE", "SPONSOR", "OTHER"]
            if guardian_rel not in valid_relations:
                errors.append({"field": "Guardian Relationship", "message": f"Must be one of {valid_relations}"})
                status = "REJECTED"

            # Academic Resolution
            row_branch_id = None
            if self.context_branch_id is not None:
                row_branch_id = self.context_branch_id
            elif branch_code:
                resolved_branch = self.repository.resolve_branch(self.tenant_id, branch_code)
                if not resolved_branch:
                    errors.append({"field": "Branch Code", "message": f"Not found: {branch_code}"})
                    status = "REJECTED"
                else:
                    row_branch_id = resolved_branch.id
                    if self.context_branch_id and row_branch_id != self.context_branch_id:
                        errors.append({"field": "Branch Code", "message": f"Unauthorized branch: {branch_code}"})
                        status = "REJECTED"
            else:
                errors.append({"field": "Branch", "message": "A target branch must be selected during upload or provided as Branch Code."})
                status = "REJECTED"

            academic_year_id = None
            if academic_year_str:
                ay = self.repository.resolve_academic_year(self.tenant_id, academic_year_str)
                if not ay:
                    errors.append({"field": "Academic Year", "message": f"Not found: {academic_year_str}"})
                    status = "REJECTED"
                else:
                    academic_year_id = ay.id

            programme_id = None
            if programme_str:
                prog = self.repository.resolve_programme(self.tenant_id, programme_str)
                if not prog:
                    errors.append({"field": "Programme / Stream", "message": f"Not found: {programme_str}"})
                    status = "REJECTED"
                else:
                    programme_id = prog.id

            batch_id = None
            section_id = None
            year_level = self._normalize_year_level(year_level_label)
            if year_level_label and not year_level:
                errors.append({"field": "Year Level", "message": "Must be First Year or Second Year."})
                status = "REJECTED"
            if academic_year_id and programme_id and row_branch_id and batch_str:
                batch = self.repository.resolve_batch(
                    self.tenant_id,
                    row_branch_id,
                    academic_year_id,
                    programme_id,
                    batch_str,
                )
                if batch is None:
                    errors.append({"field": "Batch", "message": f"Not found: {batch_str} for selected branch/year/programme."})
                    status = "REJECTED"
                elif year_level and str(batch.year_level) != year_level:
                    expected_year_label = "First Year" if str(batch.year_level) == "1" else "Second Year"
                    errors.append({"field": "Year Level", "message": f"Batch '{batch_str}' belongs to {expected_year_label}."})
                    status = "REJECTED"
                else:
                    batch_id = batch.id
                    year_level = year_level or str(batch.year_level)
                    if not section_str:
                        errors.append({"field": "Section", "message": "Required field missing."})
                        status = "REJECTED"
                    else:
                        section = self.repository.resolve_section(
                            self.tenant_id,
                            row_branch_id,
                            batch.id,
                            section_str,
                        )
                        if section is None:
                            errors.append({"field": "Section", "message": f"Not found: {section_str} for selected batch."})
                            status = "REJECTED"
                        else:
                            section_id = section.id
            elif not section_str:
                errors.append({"field": "Section", "message": "Required field missing."})
                status = "REJECTED"

            # Duplication Checks
            if admission_number and row_branch_id:
                admission_key = (str(row_branch_id), admission_number.lower())
                if admission_key in seen_admission_numbers:
                    errors.append({"field": "Admission Number", "message": f"Duplicate admission number '{admission_number}' inside this file."})
                    status = "REJECTED"
                seen_admission_numbers.add(admission_key)
            if admission_number and row_branch_id:
                if self.repository.check_admission_number_exists(self.tenant_id, row_branch_id, admission_number):
                    errors.append({"field": "Admission Number", "message": f"Admission number '{admission_number}' already exists in branch."})
                    status = "REJECTED"

            if roll_number and row_branch_id and academic_year_id:
                roll_key = (str(row_branch_id), str(academic_year_id), roll_number.lower())
                if roll_key in seen_roll_numbers:
                    errors.append({"field": "Roll No", "message": f"Duplicate roll number '{roll_number}' inside this file."})
                    status = "REJECTED"
                seen_roll_numbers.add(roll_key)
                if self.repository.check_roll_number_exists(self.tenant_id, row_branch_id, academic_year_id, roll_number):
                    errors.append({"field": "Roll No", "message": f"Roll number '{roll_number}' already exists for this branch and academic year."})
                    status = "REJECTED"

            if student_name and dob_val and guardian_mobile and self.repository.possible_student_match_exists(
                self.tenant_id, student_name, dob_val, guardian_mobile
            ):
                errors.append({"field": "Student Name", "message": "Possible existing student match with same name, date of birth and guardian mobile."})
                if status == "VALID":
                    status = "WARNING"

            if not student_mobile:
                errors.append({"field": "Student Mobile", "message": "Missing optional student mobile."})
                if status == "VALID":
                    status = "WARNING"

            normalized_data = {
                "student_name": student_name,
                "admission_number": admission_number,
                "date_of_birth": dob_val,
                "gender": gender or "UNKNOWN",
                "student_mobile": student_mobile if student_mobile else None,
                "student_email": student_email if student_email else None,
                "guardian_name": guardian_name,
                "guardian_relationship": guardian_rel,
                "guardian_mobile": guardian_mobile,
                "guardian_email": guardian_email if guardian_email else None,
                "year_level": year_level,
                "roll_number": roll_number if roll_number else None,
                "joining_date": joining_val,
                "ending_date": ending_val,
                "academic_year_id": str(academic_year_id) if academic_year_id else None,
                "programme_id": str(programme_id) if programme_id else None,
                "batch_id": str(batch_id) if batch_id else None,
                "section_id": str(section_id) if section_id else None,
                "branch_id": str(row_branch_id) if row_branch_id else None,
            }

            if status == "VALID":
                summary["valid_rows"] += 1
            elif status == "WARNING":
                summary["warning_rows"] += 1
            else:
                summary["rejected_rows"] += 1

            results.append({
                "row_number": row_idx,
                "raw_data": {k: "" if v is None else (v.isoformat() if hasattr(v, "isoformat") else str(v)) for k, v in raw_data.items()},
                "normalized_data": normalized_data,
                "validation_status": status,
                "errors": errors,
            })

        return results, summary


class FeeImportValidator:
    def __init__(self, repository: ImportRepository, tenant_id: UUID, context_branch_id: UUID | None = None) -> None:
        self.repository = repository
        self.tenant_id = tenant_id
        self.context_branch_id = context_branch_id
        self.required_columns = {
            "admission_number": ["Admission No", "Admission Number"],
            "academic_year": ["Academic Year"],
            "assigned_fee": ["Assigned Fee"],
            "payment_schedule_type": ["Payment Schedule Type"],
        }
        self.optional_columns = {
            "student_name": ["Student Name"],
            "programme": ["Programme / Stream"],
            "section": ["Section"],
            "government_scholarship": ["Government Scholarship"],
            "concession": ["Concession"],
            "notes": ["Notes"],
        }

    def _header_map(self, headers: tuple[Any, ...]) -> dict[str, int]:
        return {str(header).strip(): index for index, header in enumerate(headers) if header is not None and str(header).strip()}

    def _find_column(self, header_map: dict[str, int], aliases: list[str]) -> int | None:
        normalized = {header.lower(): index for header, index in header_map.items()}
        for alias in aliases:
            index = normalized.get(alias.lower())
            if index is not None:
                return index
        return None

    def _value(self, row_values: tuple[Any, ...], columns: dict[str, int | None], key: str) -> Any:
        index = columns.get(key)
        if index is None or index >= len(row_values):
            return None
        return row_values[index]

    def _clean_string(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, Decimal):
            return str(value.normalize())
        return str(value).strip()

    def _parse_money(self, value: Any) -> Decimal | None:
        if value in (None, ""):
            return Decimal("0")
        if isinstance(value, Decimal):
            parsed = value
        else:
            text = self._clean_string(value).replace(",", "")
            try:
                parsed = Decimal(text)
            except Exception:
                return None
        if parsed < 0:
            return None
        return parsed.quantize(Decimal("0.01"))

    def parse_and_validate(self, file_content: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active
        if not sheet:
            raise ValueError("No active sheet found in Excel file.")

        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            raise ValueError("Excel file is empty.")

        header_map = self._header_map(headers)
        columns = {
            key: self._find_column(header_map, aliases)
            for key, aliases in {**self.required_columns, **self.optional_columns}.items()
        }
        missing_headers = [
            aliases[0]
            for key, aliases in self.required_columns.items()
            if columns[key] is None
        ]
        if missing_headers:
            raise ValueError(f"Missing required columns: {', '.join(missing_headers)}")

        results = []
        summary = {"total_rows": 0, "valid_rows": 0, "warning_rows": 0, "rejected_rows": 0}
        seen_fee_keys: set[tuple[str, str]] = set()
        valid_schedule_types = {"ONE_TIME", "TERM_WISE", "INSTALLMENT_WISE", "CUSTOM"}

        for row_idx, row_values in enumerate(rows_iter, start=2):
            if not any(row_values):
                continue
            if row_idx == 2:
                first_value = self._clean_string(row_values[0] if row_values else "")
                if first_value.lower() in {"required", "optional"}:
                    continue

            summary["total_rows"] += 1
            raw_data = {header: row_values[idx] for header, idx in header_map.items()}
            errors = []
            status = "VALID"

            admission_number = self._clean_string(self._value(row_values, columns, "admission_number"))
            student_name = self._clean_string(self._value(row_values, columns, "student_name"))
            academic_year = self._clean_string(self._value(row_values, columns, "academic_year"))
            programme = self._clean_string(self._value(row_values, columns, "programme"))
            section = self._clean_string(self._value(row_values, columns, "section"))
            assigned_fee = self._parse_money(self._value(row_values, columns, "assigned_fee"))
            scholarship = self._parse_money(self._value(row_values, columns, "government_scholarship"))
            concession = self._parse_money(self._value(row_values, columns, "concession"))
            schedule_type = self._clean_string(self._value(row_values, columns, "payment_schedule_type")).upper()
            notes = self._clean_string(self._value(row_values, columns, "notes"))

            if not admission_number:
                errors.append({"field": "Admission No", "message": "Required field missing."})
                status = "REJECTED"
            if not academic_year:
                errors.append({"field": "Academic Year", "message": "Required field missing."})
                status = "REJECTED"
            if assigned_fee is None:
                errors.append({"field": "Assigned Fee", "message": "Must be a zero or positive amount."})
                status = "REJECTED"
            if scholarship is None:
                errors.append({"field": "Government Scholarship", "message": "Must be a zero or positive amount."})
                status = "REJECTED"
            if concession is None:
                errors.append({"field": "Concession", "message": "Must be a zero or positive amount."})
                status = "REJECTED"
            if schedule_type not in valid_schedule_types:
                errors.append({"field": "Payment Schedule Type", "message": "Must be one of ONE_TIME, TERM_WISE, INSTALLMENT_WISE, CUSTOM."})
                status = "REJECTED"

            assigned_amount = assigned_fee or Decimal("0")
            scholarship_amount = scholarship or Decimal("0")
            concession_amount = concession or Decimal("0")
            if scholarship_amount + concession_amount > assigned_amount:
                errors.append({"field": "Assigned Fee", "message": "Scholarship and concession cannot exceed assigned fee."})
                status = "REJECTED"

            enrollment = None
            if admission_number and academic_year:
                file_key = (admission_number.lower(), academic_year.lower())
                if file_key in seen_fee_keys:
                    errors.append({"field": "Admission No", "message": "Duplicate fee setup row for this admission number and academic year inside this file."})
                    status = "REJECTED"
                seen_fee_keys.add(file_key)

                matches = self.repository.find_fee_import_enrollments(
                    tenant_id=self.tenant_id,
                    branch_id=self.context_branch_id,
                    admission_number=admission_number,
                    academic_year=academic_year,
                )
                if len(matches) == 0:
                    errors.append({"field": "Admission No", "message": "No active enrollment found for this admission number and academic year in the current context."})
                    status = "REJECTED"
                elif len(matches) > 1:
                    errors.append({"field": "Admission No", "message": "Multiple matching enrollments found. Upload within a selected branch context."})
                    status = "REJECTED"
                else:
                    enrollment = matches[0]
                    if enrollment["fee_account_id"] is not None:
                        errors.append({"field": "Admission No", "message": "A fee account already exists for this enrollment."})
                        status = "REJECTED"
                    if student_name and student_name.lower() != (enrollment["display_name"] or enrollment["legal_name"]).lower():
                        errors.append({"field": "Student Name", "message": "Student name differs from the active enrollment. Verify before finalizing."})
                        if status == "VALID":
                            status = "WARNING"
                    if programme and enrollment["programme_name"] and programme.lower() != enrollment["programme_name"].lower():
                        errors.append({"field": "Programme / Stream", "message": "Programme differs from the active enrollment. Verify before finalizing."})
                        if status == "VALID":
                            status = "WARNING"
                    if section and enrollment["section_name"] and section.lower() != enrollment["section_name"].lower():
                        errors.append({"field": "Section", "message": "Section differs from the active enrollment. Verify before finalizing."})
                        if status == "VALID":
                            status = "WARNING"

            normalized_data = {
                "admission_number": admission_number,
                "student_name": student_name,
                "academic_year": academic_year,
                "programme": programme,
                "section": section,
                "assigned_fee_amount": str(assigned_amount),
                "scholarship_amount": str(scholarship_amount),
                "concession_amount": str(concession_amount),
                "net_payable_amount": str(assigned_amount - scholarship_amount - concession_amount),
                "payment_schedule_type": schedule_type,
                "notes": notes if notes else None,
                "enrollment_id": str(enrollment["enrollment_id"]) if enrollment else None,
                "student_id": str(enrollment["student_id"]) if enrollment else None,
                "branch_id": str(enrollment["branch_id"]) if enrollment else None,
                "academic_year_id": str(enrollment["academic_year_id"]) if enrollment else None,
            }

            if status == "VALID":
                summary["valid_rows"] += 1
            elif status == "WARNING":
                summary["warning_rows"] += 1
            else:
                summary["rejected_rows"] += 1

            results.append({
                "row_number": row_idx,
                "raw_data": {k: "" if v is None else (v.isoformat() if hasattr(v, "isoformat") else str(v)) for k, v in raw_data.items()},
                "normalized_data": normalized_data,
                "validation_status": status,
                "errors": errors,
            })

        return results, summary
