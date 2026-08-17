# ruff: noqa: E501

import hashlib
import uuid
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.modules.imports.models import ImportBatch, ImportRow
from app.modules.imports.repository import ImportRepository
from app.modules.imports.schemas import (
    ActivatePortalResponse,
    BulkActivateEligibilityResponse,
    BulkActivateSectionRequest,
    BulkActivateSectionResponse,
    ImportBatchResponse,
    ImportRowResult,
    ManualAddStudentRequest,
    PreviewResponse,
    UploadResponse,
)
from app.modules.imports.validators import FeeImportValidator, StudentImportValidator
from app.modules.students.models import Enrollment, Guardian, Student, StudentGuardianLink


class ImportService:
    def __init__(self, repository: ImportRepository, session: Session) -> None:
        self.repository = repository
        self.session = session

    def upload_student_excel(
        self, tenant_id: UUID, branch_id: UUID | None, app_user_id: UUID, file_content: bytes, filename: str, context_branch_id: UUID | None = None
    ) -> UploadResponse:
        file_hash = hashlib.sha256(file_content).hexdigest()
        idempotency_key = f"{tenant_id}-{file_hash}"

        existing_batch = self.repository.get_batch_by_idempotency_key(tenant_id, idempotency_key)
        if existing_batch is not None:
            if context_branch_id and existing_batch.branch_id and existing_batch.branch_id != context_branch_id:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for this batch.")
            return UploadResponse(
                message="This file was already uploaded. Opening the existing validation preview.",
                batch_id=existing_batch.id,
                status=existing_batch.status,
            )
        
        batch = ImportBatch(
            id=uuid.uuid4(),
            tenant_id=tenant_id,
            branch_id=branch_id,
            module_code="students",
            import_type="EXCEL_TEMPLATE",
            schema_version="V1",
            source_filename=filename,
            storage_key=f"imports/students/{file_hash}.xlsx",
            checksum=file_hash,
            idempotency_key=idempotency_key,
            status="UPLOADED",
            created_by=app_user_id,
        )
        self.repository.create_batch(batch)

        effective_branch_id = context_branch_id or branch_id
        validator = StudentImportValidator(self.repository, tenant_id, effective_branch_id)
        
        try:
            row_results, summary = validator.parse_and_validate(file_content)
        except ValueError as e:
            batch.status = "FAILED"
            self.repository.update_batch(batch)
            self.session.commit()
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e)) from e

        batch.summary = summary
        batch.status = "PREVIEW"
        self.repository.update_batch(batch)

        db_rows = []
        for res in row_results:
            db_row = ImportRow(
                id=uuid.uuid4(),
                batch_id=batch.id,
                row_number=res["row_number"],
                raw_data=res["raw_data"],
                normalized_data=res["normalized_data"],
                validation_status=res["validation_status"],
                errors=res["errors"],
                target_entity_type="Student",
            )
            db_rows.append(db_row)
            
        self.repository.create_rows(db_rows)
        self.session.commit()

        return UploadResponse(
            message="File uploaded and validated.",
            batch_id=batch.id,
            status=batch.status,
        )

    def upload_fee_excel(
        self, tenant_id: UUID, branch_id: UUID | None, app_user_id: UUID, file_content: bytes, filename: str, context_branch_id: UUID | None = None
    ) -> UploadResponse:
        file_hash = hashlib.sha256(file_content).hexdigest()
        idempotency_key = f"{tenant_id}-fees-{file_hash}"

        existing_batch = self.repository.get_batch_by_idempotency_key(tenant_id, idempotency_key)
        if existing_batch is not None:
            if existing_batch.module_code != "fees":
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="This file hash is already used by another import type.")
            if context_branch_id and existing_batch.branch_id and existing_batch.branch_id != context_branch_id:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for this batch.")
            return UploadResponse(
                message="This fee file was already uploaded. Opening the existing validation preview.",
                batch_id=existing_batch.id,
                status=existing_batch.status,
            )

        batch = ImportBatch(
            id=uuid.uuid4(),
            tenant_id=tenant_id,
            branch_id=branch_id,
            module_code="fees",
            import_type="EXCEL_TEMPLATE",
            schema_version="FEE_MVP_V1",
            source_filename=filename,
            storage_key=f"imports/fees/{file_hash}.xlsx",
            checksum=file_hash,
            idempotency_key=idempotency_key,
            status="UPLOADED",
            created_by=app_user_id,
        )
        self.repository.create_batch(batch)

        effective_branch_id = context_branch_id or branch_id
        validator = FeeImportValidator(self.repository, tenant_id, effective_branch_id)

        try:
            row_results, summary = validator.parse_and_validate(file_content)
        except ValueError as e:
            batch.status = "FAILED"
            self.repository.update_batch(batch)
            self.session.commit()
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e)) from e

        batch.summary = summary
        batch.status = "PREVIEW"
        self.repository.update_batch(batch)

        db_rows = []
        for res in row_results:
            db_rows.append(
                ImportRow(
                    id=uuid.uuid4(),
                    batch_id=batch.id,
                    row_number=res["row_number"],
                    raw_data=res["raw_data"],
                    normalized_data=res["normalized_data"],
                    validation_status=res["validation_status"],
                    errors=res["errors"],
                    proposed_action="CREATE_FEE_ACCOUNT",
                    target_entity_type="FeeAccount",
                )
            )

        self.repository.create_rows(db_rows)
        self.session.commit()

        return UploadResponse(
            message="Fee file uploaded and validated.",
            batch_id=batch.id,
            status=batch.status,
        )

    def generate_student_import_template(self, tenant_id: UUID, context_branch_id: UUID | None) -> bytes:
        headers = [
            "Admission No",
            "Student Name",
            "Gender",
            "Date Of Birth",
            "Student Mobile",
            "Student Email",
            "Academic Year",
            "Programme / Stream",
            "Section",
            "Roll No",
            "Joining Date",
            "Ending Date",
            "Guardian Name",
            "Relationship",
            "Guardian Phone",
            "Guardian Email",
            "Student Created",
        ]
        required_headers = {
            "Admission No",
            "Student Name",
            "Gender",
            "Date Of Birth",
            "Academic Year",
            "Programme / Stream",
            "Section",
            "Joining Date",
            "Guardian Name",
            "Relationship",
            "Guardian Phone",
        }

        branches = self.session.execute(
            text("""
                SELECT branch_code, display_name
                FROM sms_branches
                WHERE tenant_id = :tenant_id
                    AND status = 'ACTIVE'
                    AND (CAST(:branch_id AS uuid) IS NULL OR id = CAST(:branch_id AS uuid))
                ORDER BY display_name
            """),
            {"tenant_id": tenant_id, "branch_id": str(context_branch_id) if context_branch_id else None},
        ).fetchall()
        academic_years = self.repository.get_academic_years(tenant_id)
        programmes = self.repository.get_programmes(tenant_id)
        sections = self.session.execute(
            text("""
                SELECT
                    b.display_name AS branch_name,
                    ay.name AS academic_year,
                    ap.programme_name,
                    bt.batch_name,
                    s.section_name,
                    s.section_code
                FROM sms_sections s
                JOIN sms_batches bt
                    ON bt.tenant_id = s.tenant_id
                    AND bt.branch_id = s.branch_id
                    AND bt.id = s.batch_id
                JOIN sms_academic_years ay
                    ON ay.tenant_id = bt.tenant_id
                    AND ay.id = bt.academic_year_id
                JOIN sms_academic_programmes ap
                    ON ap.tenant_id = bt.tenant_id
                    AND ap.id = bt.programme_id
                JOIN sms_branches b
                    ON b.tenant_id = s.tenant_id
                    AND b.id = s.branch_id
                WHERE s.tenant_id = :tenant_id
                    AND s.status = 'ACTIVE'
                    AND bt.status = 'ACTIVE'
                    AND ay.status = 'ACTIVE'
                    AND ap.status = 'ACTIVE'
                    AND b.status = 'ACTIVE'
                    AND (CAST(:branch_id AS uuid) IS NULL OR s.branch_id = CAST(:branch_id AS uuid))
                ORDER BY b.display_name, ay.name, ap.programme_name, bt.batch_name, s.section_name
            """),
            {"tenant_id": tenant_id, "branch_id": str(context_branch_id) if context_branch_id else None},
        ).fetchall()

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Student Import"
        instructions = workbook.create_sheet("Instructions")
        references = workbook.create_sheet("Reference Values")

        header_fill = PatternFill("solid", fgColor="0F766E")
        required_fill = PatternFill("solid", fgColor="DCFCE7")
        locked_fill = PatternFill("solid", fgColor="F1F5F9")
        header_font = Font(color="FFFFFF", bold=True)
        note_font = Font(color="64748B", italic=True)

        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            width = max(14, min(30, len(header) + 6))
            sheet.column_dimensions[get_column_letter(index)].width = width
            if header in required_headers:
                sheet.cell(row=2, column=index, value="Required").fill = required_fill
            elif header == "Student Created":
                sheet.cell(row=2, column=index, value="System generated - leave blank").fill = locked_fill
            else:
                sheet.cell(row=2, column=index, value="Optional").fill = locked_fill
            sheet.cell(row=2, column=index).font = note_font

        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        instructions_rows = [
            ("Student Import Template", "Fill the Student Import sheet from row 3 onward. Row 1 is headers and row 2 explains required/optional fields."),
            ("Required columns", ", ".join(sorted(required_headers))),
            ("Do not fill", "Student Created is generated by the system after import."),
            ("Dates", "Use YYYY-MM-DD format for Date Of Birth, Joining Date, and Ending Date."),
            ("Admission No", "Must be unique within the branch. Existing or duplicate admission numbers will be rejected."),
            ("Student Name", "Same names are allowed, but name + DOB + guardian mobile may produce duplicate warnings in preview."),
            ("Programme / Stream", "Must match an active programme from the Reference Values sheet."),
            ("Section", "Must match an active section for the selected branch, academic year, and programme."),
            ("Relationship", "Use one of: FATHER, MOTHER, LEGAL_GUARDIAN, RELATIVE, SPONSOR, OTHER."),
            ("Preview before commit", "Upload validates rows first. Records are inserted only after preview confirmation."),
        ]
        instructions.column_dimensions["A"].width = 28
        instructions.column_dimensions["B"].width = 120
        for row_index, row in enumerate(instructions_rows, start=1):
            instructions.cell(row=row_index, column=1, value=row[0]).font = Font(bold=True)
            instructions.cell(row=row_index, column=2, value=row[1])
            instructions.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        reference_sections = [
            ("Branches", ["Branch Code", "Branch Name"], [(row.branch_code, row.display_name) for row in branches]),
            ("Academic Years", ["Academic Year"], [(year.name,) for year in academic_years]),
            ("Programmes", ["Programme / Stream"], [(programme.programme_name,) for programme in programmes]),
            ("Sections", ["Branch", "Academic Year", "Programme / Stream", "Batch", "Section", "Section Code"], [
                (row.branch_name, row.academic_year, row.programme_name, row.batch_name, row.section_name, row.section_code)
                for row in sections
            ]),
            ("Allowed Gender Values", ["Gender"], [("MALE",), ("FEMALE",), ("OTHER",)]),
            ("Allowed Relationship Values", ["Relationship"], [("FATHER",), ("MOTHER",), ("LEGAL_GUARDIAN",), ("RELATIVE",), ("SPONSOR",), ("OTHER",)]),
        ]
        current_row = 1
        for title, section_headers, rows in reference_sections:
            references.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=13)
            current_row += 1
            for col_index, header in enumerate(section_headers, start=1):
                cell = references.cell(row=current_row, column=col_index, value=header)
                cell.fill = header_fill
                cell.font = header_font
            current_row += 1
            for row in rows:
                for col_index, value in enumerate(row, start=1):
                    references.cell(row=current_row, column=col_index, value=value)
                current_row += 1
            current_row += 2
        for column in range(1, 8):
            references.column_dimensions[get_column_letter(column)].width = 28

        max_data_row = 500
        gender_validation = DataValidation(type="list", formula1='"MALE,FEMALE,OTHER"', allow_blank=False)
        relationship_validation = DataValidation(type="list", formula1='"FATHER,MOTHER,LEGAL_GUARDIAN,RELATIVE,SPONSOR,OTHER"', allow_blank=False)
        sheet.add_data_validation(gender_validation)
        sheet.add_data_validation(relationship_validation)
        gender_validation.add(f"C3:C{max_data_row}")
        relationship_validation.add(f"N3:N{max_data_row}")

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def generate_fee_import_template(self, tenant_id: UUID, context_branch_id: UUID | None) -> bytes:
        headers = [
            "Admission No",
            "Student Name",
            "Academic Year",
            "Programme / Stream",
            "Section",
            "Assigned Fee",
            "Government Scholarship",
            "Concession",
            "Payment Schedule Type",
            "Notes",
        ]
        required_headers = {"Admission No", "Academic Year", "Assigned Fee", "Payment Schedule Type"}

        eligible_enrollments = self.session.execute(
            text("""
                SELECT
                    e.admission_number,
                    s.legal_name AS student_name,
                    ay.name AS academic_year,
                    ap.programme_name,
                    sec.section_name,
                    b.display_name AS branch_name
                FROM sms_enrollments e
                JOIN sms_students s
                    ON s.tenant_id = e.tenant_id
                    AND s.id = e.student_id
                JOIN sms_academic_years ay
                    ON ay.tenant_id = e.tenant_id
                    AND ay.id = e.academic_year_id
                JOIN sms_branches b
                    ON b.tenant_id = e.tenant_id
                    AND b.id = e.branch_id
                LEFT JOIN sms_academic_programmes ap
                    ON ap.tenant_id = e.tenant_id
                    AND ap.id = e.programme_id
                LEFT JOIN sms_sections sec
                    ON sec.tenant_id = e.tenant_id
                    AND sec.branch_id = e.branch_id
                    AND sec.batch_id = e.batch_id
                    AND sec.id = e.section_id
                LEFT JOIN sms_fee_accounts fa
                    ON fa.tenant_id = e.tenant_id
                    AND fa.enrollment_id = e.id
                WHERE e.tenant_id = :tenant_id
                    AND e.status = 'ACTIVE'
                    AND e.is_current = true
                    AND s.current_status = 'ACTIVE'
                    AND fa.id IS NULL
                    AND (CAST(:branch_id AS uuid) IS NULL OR e.branch_id = CAST(:branch_id AS uuid))
                ORDER BY s.legal_name, e.admission_number
            """),
            {"tenant_id": tenant_id, "branch_id": str(context_branch_id) if context_branch_id else None},
        ).fetchall()

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Fee Import"
        instructions = workbook.create_sheet("Instructions")
        references = workbook.create_sheet("Eligible Enrollments")

        header_fill = PatternFill("solid", fgColor="0F766E")
        required_fill = PatternFill("solid", fgColor="DCFCE7")
        optional_fill = PatternFill("solid", fgColor="F1F5F9")
        header_font = Font(color="FFFFFF", bold=True)
        note_font = Font(color="64748B", italic=True)

        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(index)].width = max(16, min(34, len(header) + 6))
            note_cell = sheet.cell(row=2, column=index, value="Required" if header in required_headers else "Optional")
            note_cell.fill = required_fill if header in required_headers else optional_fill
            note_cell.font = note_font

        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        instructions_rows = [
            ("Fee Import Template", "Use this template to set up fee accounts in bulk for active student enrollments."),
            ("Required columns", ", ".join(sorted(required_headers))),
            ("Admission No", "Must match an active enrollment shown on the Eligible Enrollments sheet."),
            ("Assigned Fee", "Enter the total gross fee assigned to the student for the selected academic year."),
            ("Government Scholarship", "Enter scholarship amount if applicable. Use 0 when not applicable."),
            ("Concession", "Enter institution concession amount if applicable. Use 0 when not applicable."),
            ("Payment Schedule Type", "Use one of: ONE_TIME, TERM_WISE, INSTALLMENT_WISE, CUSTOM."),
            ("Payments and receipts", "Do not enter payments in this file. Payment posting and receipt generation are handled in the Fees module."),
            ("Upload workflow", "Upload validates rows first. Fee accounts are inserted only after preview confirmation."),
        ]
        instructions.column_dimensions["A"].width = 28
        instructions.column_dimensions["B"].width = 120
        for row_index, row in enumerate(instructions_rows, start=1):
            instructions.cell(row=row_index, column=1, value=row[0]).font = Font(bold=True)
            instructions.cell(row=row_index, column=2, value=row[1])
            instructions.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        reference_headers = ["Admission No", "Student Name", "Academic Year", "Programme / Stream", "Section", "Branch"]
        for col_index, header in enumerate(reference_headers, start=1):
            cell = references.cell(row=1, column=col_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            references.column_dimensions[get_column_letter(col_index)].width = 28
        for row_index, row in enumerate(eligible_enrollments, start=2):
            references.cell(row=row_index, column=1, value=row.admission_number)
            references.cell(row=row_index, column=2, value=row.student_name)
            references.cell(row=row_index, column=3, value=row.academic_year)
            references.cell(row=row_index, column=4, value=row.programme_name)
            references.cell(row=row_index, column=5, value=row.section_name)
            references.cell(row=row_index, column=6, value=row.branch_name)
        references.freeze_panes = "A2"
        references.auto_filter.ref = f"A1:{get_column_letter(len(reference_headers))}{max(1, len(eligible_enrollments) + 1)}"

        max_data_row = 500
        schedule_validation = DataValidation(type="list", formula1='"ONE_TIME,TERM_WISE,INSTALLMENT_WISE,CUSTOM"', allow_blank=False)
        sheet.add_data_validation(schedule_validation)
        schedule_validation.add(f"I3:I{max_data_row}")

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def get_import_preview(
        self,
        batch_id: UUID,
        tenant_id: UUID,
        context_branch_id: UUID | None,
        expected_module_code: str | None = None,
    ) -> PreviewResponse:
        batch = self.repository.get_batch(batch_id)
        if not batch or batch.tenant_id != tenant_id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Batch not found.")
        if expected_module_code is not None and batch.module_code != expected_module_code:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Batch not found.")
        if context_branch_id and batch.branch_id and batch.branch_id != context_branch_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for this batch.")
            
        rows = self.repository.get_rows(batch_id)
        
        return PreviewResponse(
            batch=ImportBatchResponse.model_validate(batch),
            rows=[ImportRowResult.model_validate(r) for r in rows]
        )

    def correct_import_row(
        self,
        batch_id: UUID,
        row_id: UUID,
        tenant_id: UUID,
        raw_data: dict[str, Any],
        context_branch_id: UUID | None,
    ) -> PreviewResponse:
        batch = self.repository.get_batch(batch_id)
        if not batch or batch.tenant_id != tenant_id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Batch not found.")
        if context_branch_id and batch.branch_id and batch.branch_id != context_branch_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for this batch.")
        if batch.status not in ["PREVIEW", "FAILED"]:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only preview batches can be corrected.")

        rows = list(self.repository.get_rows(batch_id))
        target_row = next((row for row in rows if row.id == row_id), None)
        if target_row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import row not found.")

        target_row.raw_data = raw_data

        headers = [
            "Admission No",
            "Student Name",
            "Gender",
            "Date Of Birth",
            "Student Mobile",
            "Student Email",
            "Academic Year",
            "Programme / Stream",
            "Section",
            "Roll No",
            "Joining Date",
            "Ending Date",
            "Guardian Name",
            "Relationship",
            "Guardian Phone",
            "Guardian Email",
            "Student Created",
        ]
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.append(headers)
        for row in rows:
            source = row.raw_data or {}
            sheet.append([source.get(header, "") for header in headers])
        output = BytesIO()
        workbook.save(output)

        validator = StudentImportValidator(self.repository, tenant_id, context_branch_id or batch.branch_id)
        row_results, summary = validator.parse_and_validate(output.getvalue())

        for import_row, result in zip(rows, row_results, strict=False):
            import_row.raw_data = result["raw_data"]
            import_row.normalized_data = result["normalized_data"]
            import_row.validation_status = result["validation_status"]
            import_row.errors = result["errors"]
            self.repository.update_batch(batch)

        batch.summary = summary
        batch.status = "PREVIEW"
        self.session.commit()
        return self.get_import_preview(batch_id, tenant_id, context_branch_id, expected_module_code="students")

    def correct_fee_import_row(
        self,
        batch_id: UUID,
        row_id: UUID,
        tenant_id: UUID,
        raw_data: dict[str, Any],
        context_branch_id: UUID | None,
    ) -> PreviewResponse:
        batch = self.repository.get_batch(batch_id)
        if not batch or batch.tenant_id != tenant_id or batch.module_code != "fees":
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Fee import batch not found.")
        if context_branch_id and batch.branch_id and batch.branch_id != context_branch_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for this batch.")
        if batch.status not in ["PREVIEW", "FAILED"]:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only preview batches can be corrected.")

        rows = list(self.repository.get_rows(batch_id))
        target_row = next((row for row in rows if row.id == row_id), None)
        if target_row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import row not found.")

        target_row.raw_data = raw_data
        headers = [
            "Admission No",
            "Student Name",
            "Academic Year",
            "Programme / Stream",
            "Section",
            "Assigned Fee",
            "Government Scholarship",
            "Concession",
            "Payment Schedule Type",
            "Notes",
        ]
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.append(headers)
        for row in rows:
            source = row.raw_data or {}
            sheet.append([source.get(header, "") for header in headers])
        output = BytesIO()
        workbook.save(output)

        validator = FeeImportValidator(self.repository, tenant_id, context_branch_id or batch.branch_id)
        row_results, summary = validator.parse_and_validate(output.getvalue())

        for import_row, result in zip(rows, row_results, strict=False):
            import_row.raw_data = result["raw_data"]
            import_row.normalized_data = result["normalized_data"]
            import_row.validation_status = result["validation_status"]
            import_row.errors = result["errors"]
            self.repository.update_batch(batch)

        batch.summary = summary
        batch.status = "PREVIEW"
        self.session.commit()
        return self.get_import_preview(batch_id, tenant_id, context_branch_id, expected_module_code="fees")

    def commit_student_import(self, batch_id: UUID, tenant_id: UUID, app_user_id: UUID, context_branch_id: UUID | None) -> dict[str, Any]:
        batch = self.repository.get_batch(batch_id)
        if not batch or batch.tenant_id != tenant_id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Batch not found.")
        if context_branch_id and batch.branch_id and batch.branch_id != context_branch_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for this batch.")
            
        if batch.status not in ["PREVIEW", "SUBMITTED"]:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Batch is not ready to commit.")
            
        rows = self.repository.get_rows(batch_id)
        
        # Check if any rejected rows exist
        if any(r.validation_status == "REJECTED" for r in rows):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot commit batch with REJECTED rows.")

        try:
            for row in rows:
                if row.validation_status in ["VALID", "WARNING"]:
                    data = row.normalized_data
                    if not data:
                        continue
                    student_id = uuid.uuid4()
                    enrollment_id = uuid.uuid4()
                    
                    student = Student(
                        id=student_id,
                        tenant_id=tenant_id,
                        student_number=self.repository.generate_student_number(tenant_id),
                        legal_name=data["student_name"],
                        display_name=data["student_name"],
                        date_of_birth=datetime.fromisoformat(data["date_of_birth"]).date() if data.get("date_of_birth") else None,
                        gender=data["gender"],
                        student_mobile=data.get("student_mobile"),
                        student_email=data.get("student_email"),
                        current_status="ACTIVE",
                        source_type="IMPORT",
                        created_by=app_user_id,
                    )
                    self.session.add(student)
                    
                    # Match or Create Guardian
                    from sqlalchemy import select
                    stmt = select(Guardian).where(Guardian.tenant_id == tenant_id, Guardian.mobile == data["guardian_mobile"])
                    guardian = self.session.scalars(stmt).first()
                    if not guardian:
                        guardian_id = uuid.uuid4()
                        guardian = Guardian(
                            id=guardian_id,
                            tenant_id=tenant_id,
                            full_name=data["guardian_name"],
                            mobile=data["guardian_mobile"],
                            email=data.get("guardian_email"),
                            verification_status="UNVERIFIED",
                            status="ACTIVE",
                            created_by=app_user_id,
                        )
                        self.session.add(guardian)
                    else:
                        guardian_id = guardian.id
                    
                    link = StudentGuardianLink(
                        id=uuid.uuid4(),
                        tenant_id=tenant_id,
                        student_id=student_id,
                        guardian_id=guardian_id,
                        relationship_type=data["guardian_relationship"],
                        is_primary=True,
                        verification_status="PENDING",
                        status="ACTIVE",
                        created_by=app_user_id,
                    )
                    self.session.add(link)
                    
                    self.session.flush() # Force insert of Student and Guardian before Enrollment
                    
                    enrollment = Enrollment(
                        id=enrollment_id,
                        tenant_id=tenant_id,
                        student_id=student_id,
                        branch_id=UUID(data["branch_id"]) if data.get("branch_id") else None,
                        academic_year_id=UUID(data["academic_year_id"]) if data.get("academic_year_id") else None,
                        programme_id=UUID(data["programme_id"]) if data.get("programme_id") else None,
                        batch_id=UUID(data["batch_id"]) if data.get("batch_id") else None,
                        section_id=UUID(data["section_id"]) if data.get("section_id") else None,
                        admission_number=data["admission_number"],
                        roll_number=data.get("roll_number"),
                        year_level=data["year_level"],
                        status="ACTIVE",
                        joining_date=datetime.fromisoformat(data["joining_date"]).date() if data.get("joining_date") else None,
                        ending_date=datetime.fromisoformat(data["ending_date"]).date() if data.get("ending_date") else None,
                        is_current=True,
                        source_type="IMPORT",
                        created_by=app_user_id,
                    )
                    self.session.add(enrollment)
                    
                    row.target_entity_id = student_id

            batch.status = "COMMITTED"
            batch.committed_at = datetime.now(UTC)
            self.session.commit()
            
            return {"message": "Batch committed successfully", "batch_id": batch.id}
        except Exception as e:
            self.session.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to commit batch: {str(e)}",
            ) from e

    def commit_fee_import(self, batch_id: UUID, tenant_id: UUID, app_user_id: UUID, context_branch_id: UUID | None) -> dict[str, Any]:
        from app.modules.fees.repository import FeeRepository

        batch = self.repository.get_batch(batch_id)
        if not batch or batch.tenant_id != tenant_id or batch.module_code != "fees":
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Fee import batch not found.")
        if context_branch_id and batch.branch_id and batch.branch_id != context_branch_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for this batch.")
        if batch.status not in ["PREVIEW", "SUBMITTED"]:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Batch is not ready to commit.")

        rows = self.repository.get_rows(batch_id)
        if any(row.validation_status == "REJECTED" for row in rows):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot commit batch with REJECTED rows.")

        fee_repository = FeeRepository(self.session)
        committed_count = 0
        try:
            for row in rows:
                if row.validation_status not in ["VALID", "WARNING"] or not row.normalized_data:
                    continue

                data = row.normalized_data
                enrollment = fee_repository.get_active_enrollment_for_fee_setup(
                    tenant_id=tenant_id,
                    branch_id=context_branch_id or batch.branch_id,
                    enrollment_id=UUID(data["enrollment_id"]),
                )
                if enrollment is None:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Row {row.row_number}: Active enrollment no longer exists.")
                if enrollment["fee_account_id"] is not None:
                    raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"Row {row.row_number}: Fee account already exists.")

                assigned_fee_amount = Decimal(data["assigned_fee_amount"])
                scholarship_amount = Decimal(data["scholarship_amount"])
                concession_amount = Decimal(data["concession_amount"])
                net_payable = assigned_fee_amount - scholarship_amount - concession_amount
                account_id = fee_repository.create_fee_account(
                    {
                        "id": uuid.uuid4(),
                        "tenant_id": tenant_id,
                        "branch_id": enrollment["branch_id"],
                        "student_id": enrollment["student_id"],
                        "enrollment_id": enrollment["enrollment_id"],
                        "academic_year_id": enrollment["academic_year_id"],
                        "assigned_fee_amount": assigned_fee_amount,
                        "scholarship_amount": scholarship_amount,
                        "concession_amount": concession_amount,
                        "net_payable_amount": net_payable,
                        "total_paid_amount": Decimal("0"),
                        "total_adjusted_amount": Decimal("0"),
                        "total_reversed_amount": Decimal("0"),
                        "outstanding_amount": net_payable,
                        "payment_schedule_type": data["payment_schedule_type"],
                        "payment_schedule": [],
                        "status": "PAID" if net_payable == 0 else "ACTIVE",
                        "created_by": app_user_id,
                        "metadata": {"import_batch_id": str(batch.id), "source": "FEE_IMPORT"},
                    }
                )
                self._create_fee_import_ledger_entries(
                    fee_repository=fee_repository,
                    tenant_id=tenant_id,
                    account_id=account_id,
                    enrollment=enrollment,
                    app_user_id=app_user_id,
                    assigned_fee_amount=assigned_fee_amount,
                    scholarship_amount=scholarship_amount,
                    concession_amount=concession_amount,
                    notes=data.get("notes"),
                    batch_id=batch.id,
                )
                row.target_entity_id = account_id
                committed_count += 1

            batch.status = "COMMITTED"
            batch.committed_at = datetime.now(UTC)
            batch.summary = {**(batch.summary or {}), "committed_rows": committed_count}
            self.session.commit()
            return {"message": "Fee import committed successfully", "batch_id": batch.id, "committed_rows": committed_count}
        except HTTPException:
            self.session.rollback()
            raise
        except Exception as e:
            self.session.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to commit fee batch: {str(e)}",
            ) from e

    def _create_fee_import_ledger_entries(
        self,
        *,
        fee_repository: Any,
        tenant_id: UUID,
        account_id: UUID,
        enrollment: dict[str, object],
        app_user_id: UUID,
        assigned_fee_amount: Decimal,
        scholarship_amount: Decimal,
        concession_amount: Decimal,
        notes: str | None,
        batch_id: UUID,
    ) -> None:
        common = {
            "tenant_id": tenant_id,
            "branch_id": enrollment["branch_id"],
            "fee_account_id": account_id,
            "student_id": enrollment["student_id"],
            "enrollment_id": enrollment["enrollment_id"],
            "academic_year_id": enrollment["academic_year_id"],
            "posted_by": app_user_id,
            "created_by": app_user_id,
            "status": "POSTED",
            "metadata": {"import_batch_id": str(batch_id), "source": "FEE_IMPORT"},
        }
        if assigned_fee_amount > 0:
            fee_repository.create_ledger_entry(
                {
                    **common,
                    "entry_type": "FEE_ASSIGNED",
                    "balance_effect": "INCREASE",
                    "amount": assigned_fee_amount,
                    "notes": notes or "Initial fee account setup from fee import.",
                }
            )
        if scholarship_amount > 0:
            fee_repository.create_ledger_entry(
                {
                    **common,
                    "entry_type": "GOVERNMENT_SCHOLARSHIP",
                    "balance_effect": "DECREASE",
                    "amount": scholarship_amount,
                    "notes": notes or "Scholarship recorded from fee import.",
                }
            )
        if concession_amount > 0:
            fee_repository.create_ledger_entry(
                {
                    **common,
                    "entry_type": "CONCESSION",
                    "balance_effect": "DECREASE",
                    "amount": concession_amount,
                    "notes": notes or "Concession recorded from fee import.",
                }
            )

    def create_manual_student(
        self, tenant_id: UUID, branch_id: UUID, app_user_id: UUID, payload: "ManualAddStudentRequest"
    ) -> dict[str, Any]:
        # 1. Scope Validation & Academic Validation
        from app.modules.academic_structure.models import Batch, Section
        section = self.session.get(Section, payload.section_id)
        if not section or section.batch_id != payload.batch_id or section.tenant_id != tenant_id or section.branch_id != branch_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid academic hierarchy: Section does not belong to the selected Batch/Branch.")
            
        batch_model = self.session.get(Batch, payload.batch_id)
        if not batch_model or batch_model.programme_id != payload.programme_id or batch_model.academic_year_id != payload.academic_year_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid academic hierarchy: Batch does not belong to the selected Programme/Academic Year.")
        
        # 2. Duplicate Checks
        if self.repository.check_admission_number_exists(tenant_id, branch_id, payload.admission_number):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Admission number already exists in this branch.")
            
        # 3. Generate Student Number
        student_number = self.repository.generate_student_number(tenant_id)
            
        try:
            # 4. Create Student
            student_id = uuid.uuid4()
            student = Student(
                id=student_id,
                tenant_id=tenant_id,
                student_number=student_number,
                legal_name=payload.student_name,
                display_name=payload.student_name,
                date_of_birth=payload.date_of_birth,
                gender=payload.gender,
                current_status="ACTIVE",
                source_type="MANUAL",
                created_by=app_user_id,
            )
            self.session.add(student)
            
            # 5. Match/Create Guardian
            # Lookup guardian by mobile in tenant
            # Here we assume a repo method get_guardian_by_mobile exists, or we just create a new one for MVP
            from sqlalchemy import select
            stmt = select(Guardian).where(Guardian.tenant_id == tenant_id, Guardian.mobile == payload.guardian_mobile)
            guardian = self.session.scalars(stmt).first()
            if not guardian:
                guardian_id = uuid.uuid4()
                guardian = Guardian(
                    id=guardian_id,
                    tenant_id=tenant_id,
                    full_name=payload.guardian_name,
                    mobile=payload.guardian_mobile,
                    email=payload.guardian_email,
                    verification_status="UNVERIFIED",
                    status="ACTIVE",
                    created_by=app_user_id,
                )
                self.session.add(guardian)
            else:
                guardian_id = guardian.id
                
            # 6. Create Guardian Link
            link = StudentGuardianLink(
                id=uuid.uuid4(),
                tenant_id=tenant_id,
                student_id=student_id,
                guardian_id=guardian_id,
                relationship_type=payload.relationship_type,
                is_primary=True,
                verification_status="PENDING",
                status="ACTIVE",
                created_by=app_user_id,
            )
            self.session.add(link)
            
            self.session.flush() # Force insert before Enrollment
            
            # 7. Create Enrolment
            enrollment_id = uuid.uuid4()
            enrollment = Enrollment(
                id=enrollment_id,
                tenant_id=tenant_id,
                student_id=student_id,
                branch_id=branch_id,
                academic_year_id=payload.academic_year_id,
                programme_id=payload.programme_id,
                batch_id=payload.batch_id,
                section_id=payload.section_id,
                admission_number=payload.admission_number,
                roll_number=payload.roll_number if payload.roll_number else None,
                year_level=payload.year_level,
                status="ACTIVE",
                joining_date=datetime.now(UTC).date(),
                is_current=True,
                source_type="MANUAL",
                created_by=app_user_id,
            )
            self.session.add(enrollment)
            
            # 8. Commit
            self.session.commit()
            
            return {
                "student_id": student_id,
                "student_number": student_number,
                "guardian_id": guardian_id,
                "enrollment_id": enrollment_id
            }
        except Exception as e:
            self.session.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Transaction failed: {str(e)}",
            ) from e

    def activate_parent_portal(self, tenant_id: UUID, guardian_id: UUID, app_user_id: UUID) -> ActivatePortalResponse:
        """
        Orchestration contract for single guardian activation.
        Currently blocked by external dependency.
        """
        guardian = self.repository.get_guardian_with_links(guardian_id, tenant_id)
        if not guardian:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Guardian not found or inactive")
        
        # If already active, just return success
        if guardian.portal_user_id:
            return ActivatePortalResponse(
                guardian_id=guardian.id,
                portal_user_id=guardian.portal_user_id,
                status="ALREADY_ACTIVE",
                message="Parent Portal is already active for this guardian."
            )
            
        # External Dependency Block
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED, 
            detail="External Dependency — Parent Account Provisioning Service"
        )

    def get_bulk_activation_eligibility(self, tenant_id: UUID, payload: BulkActivateSectionRequest) -> BulkActivateEligibilityResponse:
        """
        Calculates eligibility preview for bulk activation.
        """
        guardians = self.repository.get_guardians_for_section(
            tenant_id=tenant_id,
            branch_id=payload.branch_id,
            academic_year_id=payload.academic_year_id,
            programme_id=payload.programme_id,
            batch_id=payload.batch_id,
            section_id=payload.section_id
        )
        
        total_students = 0 # Handled externally or mock
        unique_guardians = len(guardians)
        already_active_count = sum(1 for g in guardians if g.portal_user_id is not None)
        missing_contact_count = sum(1 for g in guardians if not g.mobile and not g.email)
        eligible_count = unique_guardians - already_active_count - missing_contact_count
        
        return BulkActivateEligibilityResponse(
            total_students=total_students,
            unique_guardians=unique_guardians,
            eligible_count=eligible_count,
            already_active_count=already_active_count,
            missing_contact_count=missing_contact_count
        )

    def bulk_activate_parent_portal(self, tenant_id: UUID, payload: BulkActivateSectionRequest, app_user_id: UUID) -> BulkActivateSectionResponse:
        """
        Orchestration contract for bulk activation by section.
        Currently blocked by external dependency.
        """
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED, 
            detail="External Dependency — Parent Account Provisioning Service"
        )
